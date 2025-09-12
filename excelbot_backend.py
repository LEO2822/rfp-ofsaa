from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import pandas as pd
import io
import os
import uuid
import asyncio
import textwrap
from datetime import datetime
from dotenv import load_dotenv
import re
import json
import time
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openai import AsyncOpenAI
import mcp
from mcp.client.streamable_http import streamablehttp_client
from mcp.shared.exceptions import McpError
from functools import lru_cache
import hashlib
from concurrent.futures import ThreadPoolExecutor
from loguru import logger

# Load environment variables
load_dotenv()

# OpenAI API key
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    logger.error("Missing OPENAI_API_KEY in .env file")
    raise EnvironmentError("Missing OPENAI_API_KEY in .env file")

# Initialize OpenAI client with optimized settings
client = AsyncOpenAI(
    api_key=OPENAI_API_KEY,
    max_retries=3,
    timeout=30.0
)

# Smithery.ai Exa Search configuration
SMITHERY_API_KEY = os.getenv("SMITHERY_API_KEY")
SMITHERY_PROFILE = os.getenv("SMITHERY_PROFILE")
if not SMITHERY_API_KEY or not SMITHERY_PROFILE:
    logger.error("Missing SMITHERY_API_KEY or SMITHERY_PROFILE in .env file")
    raise EnvironmentError("Missing SMITHERY_API_KEY or SMITHERY_PROFILE in .env file")

EXA_URL = f"https://server.smithery.ai/exa/mcp?api_key={SMITHERY_API_KEY}&profile={SMITHERY_PROFILE}"
EXA_TOOL = "web_search_exa"

# Optimized model selection
MODEL_NAME = "gpt-4o-mini"

# Performance configurations
MAX_CONCURRENT_REQUESTS = 10
BATCH_SIZE = 5
RATE_LIMIT_DELAY = 0.1
MAX_RETRIES = 3
CACHE_SIZE = 1000

app = FastAPI(title="Enhanced AI Excel Processor - Unbiased Analysis")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3001"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class ProcessRequest(BaseModel):
    input_columns: List[str]
    output_columns: List[str]
    filename: str
    user_prompt: str = ""

# Cache for search results
search_cache: Dict[str, Dict[str, Any]] = {}
semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)

def get_cache_key(query: str) -> str:
    """Generate cache key for search queries"""
    return hashlib.md5(query.lower().strip().encode()).hexdigest()

def analyze_source_type(url: str) -> str:
    """Categorize source types for better analysis"""
    url_lower = url.lower()
    if 'oracle.com' in url_lower:
        if 'docs.oracle.com' in url_lower:
            return "Official Oracle Documentation"
        elif 'support.oracle.com' in url_lower:
            return "Oracle Support Resources"
        elif 'blogs.oracle.com' in url_lower:
            return "Oracle Technical Blogs"
        else:
            return "Oracle Official Website"
    elif any(domain in url_lower for domain in ['github.com', 'stackoverflow.com']):
        return "Developer Community Resources"
    elif any(domain in url_lower for domain in ['.edu', 'research', 'academic']):
        return "Academic/Research Resources"
    elif any(domain in url_lower for domain in ['techcrunch', 'zdnet', 'computerweekly', 'itworld', 'finextra']):
        return "Technology News/Analysis"
    elif 'banking' in url_lower or 'financial' in url_lower:
        return "Banking/Financial Industry Resources"
    else:
        return "Industry/Technical Articles"

# ===============================
# Enhanced Exa Web Search with Source Tracking
# ===============================
async def exa_search(query: str, row_index: Optional[int] = None) -> Dict[str, Any]:
    """Enhanced web search with comprehensive source tracking and result quality assessment"""
    cache_key = get_cache_key(query)
    
    # Check cache first
    if cache_key in search_cache:
        if row_index is not None:
            logger.debug(f"Row {row_index + 1} - Using cached search for: '{query[:50]}...'")
        return search_cache[cache_key]
    
    async with semaphore:
        try:
            await asyncio.sleep(RATE_LIMIT_DELAY)
            
            async with streamablehttp_client(EXA_URL) as (r, w, _):
                async with mcp.ClientSession(r, w) as s:
                    await s.initialize()
                    tools = [t.name for t in (await s.list_tools()).tools]

                    if EXA_TOOL not in tools:
                        result = {
                            "content": f"Error: Exa search tool not available. Found: {tools}",
                            "sources": [],
                            "source_types": [],
                            "evidence_strength": "None",
                            "oracle_sources": 0,
                            "community_sources": 0
                        }
                        search_cache[cache_key] = result
                        return result

                    try:
                        res = await s.call_tool(EXA_TOOL, {"query": query})
                    except McpError as e:
                        result = {
                            "content": f"Exa search error: {str(e)}",
                            "sources": [],
                            "source_types": [],
                            "evidence_strength": "None",
                            "oracle_sources": 0,
                            "community_sources": 0
                        }
                        search_cache[cache_key] = result
                        return result

                    data = None
                    if hasattr(res, 'content') and res.content:
                        data = res.content
                    elif hasattr(res, 'output') and res.output:
                        data = res.output
                    elif hasattr(res, 'result') and res.result:
                        data = res.result

                    if not data:
                        result = {
                            "content": "No web search results found.",
                            "sources": [],
                            "source_types": [],
                            "evidence_strength": "None",
                            "oracle_sources": 0,
                            "community_sources": 0
                        }
                        search_cache[cache_key] = result
                        return result

                    results = []
                    urls_found = []
                    source_types = []
                    
                    def process_search_data(data) -> List[str]:
                        processed_results = []
                        
                        if isinstance(data, list):
                            for i, item in enumerate(data[:5], 1):
                                if hasattr(item, 'text'):
                                    text_content = item.text
                                    if text_content.strip().startswith('{'):
                                        try:
                                            json_data = json.loads(text_content)
                                            if 'results' in json_data:
                                                for j, result in enumerate(json_data['results'][:5], 1):
                                                    title = result.get('title', 'No title')
                                                    snippet = result.get('snippet') or result.get('text', '')
                                                    url = result.get('url', '')
                                                    if url:
                                                        urls_found.append(url)
                                                        source_types.append(analyze_source_type(url))
                                                    formatted_result = f"{j}. {title}"
                                                    if snippet:
                                                        formatted_result += f"\n   {snippet[:500]}..."
                                                    processed_results.append(formatted_result)
                                        except json.JSONDecodeError:
                                            processed_results.append(f"{i}. {text_content[:400]}")
                                elif isinstance(item, dict):
                                    title = item.get("title", "No title")
                                    snippet = item.get("snippet") or item.get("text") or ""
                                    url = item.get("url", "")
                                    if url:
                                        urls_found.append(url)
                                        source_types.append(analyze_source_type(url))
                                    result = f"{i}. {title}"
                                    if snippet:
                                        result += f"\n   {snippet[:500]}..."
                                    processed_results.append(result)
                                else:
                                    processed_results.append(f"{i}. {str(item)[:400]}")

                        elif hasattr(data, 'text') or isinstance(data, str):
                            text_content = data.text if hasattr(data, 'text') else str(data)
                            if text_content.strip().startswith('{'):
                                try:
                                    json_data = json.loads(text_content)
                                    if 'results' in json_data:
                                        for i, result in enumerate(json_data['results'][:5], 1):
                                            title = result.get('title', 'No title')
                                            snippet = result.get('snippet') or result.get('text', '')
                                            url = result.get('url', '')
                                            if url:
                                                urls_found.append(url)
                                                source_types.append(analyze_source_type(url))
                                            formatted_result = f"{i}. {title}"
                                            if snippet:
                                                formatted_result += f"\n   {snippet[:500]}..."
                                            processed_results.append(formatted_result)
                                except json.JSONDecodeError:
                                    processed_results.append(f"1. {text_content[:600]}")
                            else:
                                processed_results.append(f"1. {text_content[:600]}")
                        
                        return processed_results
                    
                    results = process_search_data(data)
                    
                    # Analyze source quality
                    oracle_sources = len([url for url in urls_found if 'oracle.com' in url.lower()])
                    community_sources = len([url for url in urls_found if 'oracle.com' not in url.lower()])
                    
                    # Determine evidence strength
                    evidence_strength = "None"
                    if oracle_sources >= 2:
                        evidence_strength = "High"
                    elif oracle_sources >= 1 or (community_sources >= 3):
                        evidence_strength = "Moderate" 
                    elif urls_found:
                        evidence_strength = "Limited"
                    
                    # Enhanced terminal display
                    if row_index is not None:
                        print(f"\n{'='*120}")
                        print(f"🔍 ROW {row_index + 1} - COMPREHENSIVE SEARCH ANALYSIS")
                        print(f"{'='*120}")
                        print(f"📝 Search Query: '{query}'")
                        print(f"🎯 Cache Status: {'HIT' if cache_key in search_cache else 'MISS'}")
                        print(f"📊 Results Quality: {len(results)} detailed results found")
                        print(f"🏆 Evidence Strength: {evidence_strength}")
                        
                        if urls_found:
                            print(f"\n🔗 REFERENCE SOURCES FOUND ({len(urls_found)} total):")
                            print(f"   📋 Oracle Official Sources: {oracle_sources}")
                            print(f"   🌐 Community/Industry Sources: {community_sources}")
                            
                            for i, (url, source_type) in enumerate(zip(urls_found[:8], source_types[:8]), 1):
                                print(f"   {i}. [{source_type}]")
                                print(f"      {url}")
                        
                        if source_types:
                            unique_source_types = list(set(source_types))
                            print(f"\n📚 SOURCE CATEGORIES COVERED ({len(unique_source_types)}):")
                            for st in unique_source_types:
                                count = source_types.count(st)
                                print(f"   • {st} ({count} sources)")
                        
                        logger.info("\n📋 DETAILED SEARCH RESULTS:")
                        if results:
                            for i, result in enumerate(results[:4], 1):
                                logger.info(f"\n   📄 Result {i}:")
                                lines = result.split('\n')
                                for line in lines:
                                    logger.info(f"      {textwrap.fill(line, width=100, subsequent_indent='      ')}")
                        else:
                            logger.warning("   ❌ No detailed results found")
                        
                        logger.info("\n🎯 ASSESSMENT GUIDANCE:")
                        if evidence_strength == "High":
                            logger.info("   ✅ Strong evidence available - should support decisive Yes/No responses")
                        elif evidence_strength == "Moderate":
                            logger.info("   ⚡ Moderate evidence - evaluate specific findings for appropriate response")
                        elif evidence_strength == "Limited":
                            logger.warning("   ⚠️  Limited evidence - may warrant 'Partially' or 'Not found' response")
                        else:
                            logger.warning("   ❌ No evidence - should result in 'Not found' response")
                        
                        logger.info(f"{'='*120}\n")

                    final_content = "\n\n".join(results) if results else "No meaningful search results found."
                    
                    result = {
                        "content": final_content,
                        "sources": urls_found,
                        "source_types": source_types,
                        "evidence_strength": evidence_strength,
                        "oracle_sources": oracle_sources,
                        "community_sources": community_sources
                    }
                    
                    if len(search_cache) < CACHE_SIZE:
                        search_cache[cache_key] = result
                    
                    return result

        except Exception as e:
            error_msg = f"Web search failed: {str(e)}"
            if row_index is not None:
                logger.error(f"\n❌ ROW {row_index + 1} - SEARCH ERROR")
                logger.error(f"{'='*80}")
                logger.error(f"Query: '{query}'")
                logger.error(f"Error: {error_msg}")
                logger.error(f"{'='*80}\n")
            logger.exception(f"Search error for query: {query}")
            
            result = {
                "content": error_msg,
                "sources": [],
                "source_types": [],
                "evidence_strength": "Error",
                "oracle_sources": 0,
                "community_sources": 0
            }
            search_cache[cache_key] = result
            return result

# ===============================
# Optimized OpenAI Call
# ===============================
async def call_openai(messages: list, max_tokens: int = 1200, retry_count: int = 0) -> str:
    """Optimized OpenAI API call with retry logic"""
    try:
        response = await client.chat.completions.create(
            model=MODEL_NAME,
            messages=messages,
            max_tokens=max_tokens,
            temperature=0.1,
            frequency_penalty=0.2,
            presence_penalty=0.1
        )
        return response.choices[0].message.content
    except Exception as e:
        if retry_count < MAX_RETRIES:
            logger.warning(f"OpenAI API retry {retry_count + 1}/{MAX_RETRIES}: {str(e)}")
            await asyncio.sleep(2 ** retry_count)
            return await call_openai(messages, max_tokens, retry_count + 1)
        else:
            logger.error(f"OpenAI API call failed after {MAX_RETRIES} retries: {str(e)}")
            return f"OpenAI API call failed after {MAX_RETRIES} retries: {str(e)}"

# ===============================
# Upload Excel
# ===============================
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        if len(contents) > 100 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File exceeds 100MB limit")
        if not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are supported")

        df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_filename = f"temp_{timestamp}_{file.filename}"

        if not hasattr(app.state, 'temp_files'):
            app.state.temp_files = {}
        app.state.temp_files[temp_filename] = contents

        return {
            "filename": temp_filename,
            "columns": df.columns.tolist(),
            "row_count": len(df),
            "original_filename": file.filename
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload error: {str(e)}")

# ===============================
# Enhanced Column Value Extraction
# ===============================
def extract_column_values(text: str, output_cols: List[str], search_info: Dict[str, Any]) -> Dict[str, str]:
    """Enhanced column value extraction with response-specific formatting (NO LINKS)"""
    column_mapping = {
        "RESPONSE": "TENDERER'S RESPONSE",
        "REMARK": "TENDERER'S REMARK", 
        "COMPLIANCE": "TENDERER'S RESPONSE",
        "COMMENT": "TENDERER'S REMARK",
        "VENDOR RESPONSE": "VENDOR RESPONSE",
        "VENDOR REMARKS": "VENDOR REMARKS",
        "VENDOR COMMENTS": "VENDOR REMARKS",
        "ANSWER": "ANSWER",
        "NOTES": "NOTES",
        "TENDERER'S RESPONSE": "TENDERER'S RESPONSE",
        "TENDERER'S REMARK": "TENDERER'S REMARK"
    }
    
    results = {col: "" for col in output_cols}
    current_col = None
    buffer = []
    explanation = ""
    
    lines = text.splitlines()
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line.upper().startswith("EXPLANATION:"):
            explanation = line.split(":", 1)[1].strip() if ":" in line else ""
            continue
        
        match = re.match(r"^([A-Za-z_'\s]+)\s*:\s*(.*)", line, re.IGNORECASE)
        if match:
            if current_col and current_col in results:
                results[current_col] = "\n".join(buffer).strip()
            
            raw_col = match.group(1).strip().upper()
            current_col = column_mapping.get(raw_col, raw_col)
            
            if current_col not in output_cols:
                for output_col in output_cols:
                    if any(key in raw_col for key in ["RESPONSE", "ANSWER", "COMPLIANCE"]) and any(key in output_col for key in ["RESPONSE", "ANSWER", "COMPLIANCE"]):
                        current_col = output_col
                        break
                    elif any(key in raw_col for key in ["REMARK", "COMMENT", "NOTES"]) and any(key in output_col for key in ["REMARK", "COMMENT", "NOTES"]):
                        current_col = output_col
                        break
                
                if current_col not in output_cols:
                    current_col = None
                    continue
            
            buffer = [match.group(2).strip()]
        elif current_col:
            buffer.append(line)

    if current_col and current_col in results:
        results[current_col] = "\n".join(buffer).strip()

    # Fill empty columns with appropriate content
    for col in output_cols:
        if not results[col].strip():
            if any(keyword in col.upper() for keyword in ["RESPONSE", "ANSWER", "COMPLIANCE"]):
                results[col] = "Not found"
            elif any(keyword in col.upper() for keyword in ["REMARK", "COMMENT", "NOTES"]):
                results[col] = "Based on comprehensive analysis of available Oracle documentation and industry resources, specific information regarding this requirement could not be definitively established."

    # Find response and remark columns for conditional formatting
    response_col = None
    remark_col = None
    
    for col in output_cols:
        if any(keyword in col.upper() for keyword in ["RESPONSE", "ANSWER", "COMPLIANCE"]):
            response_col = col
        elif any(keyword in col.upper() for keyword in ["REMARK", "COMMENT", "NOTES"]):
            remark_col = col
    
    # Apply conditional formatting based on response value (NO REFERENCE SOURCES)
    if response_col and remark_col and results[response_col]:
        response_value = results[response_col].strip().lower()
        
        if response_value == "yes":
            # YES: explanation only (no links)
            if explanation:
                results[remark_col] = explanation
            else:
                results[remark_col] = "Oracle FLEXCUBE provides the required functionality as part of its core banking capabilities."
                
        elif response_value == "partially":
            # PARTIALLY: explanation only (no links)
            if explanation:
                results[remark_col] = explanation
            else:
                results[remark_col] = "Oracle FLEXCUBE provides partial support for this requirement with some limitations or additional configuration needed."
                
        elif response_value == "no":
            # NO: explanation only
            if explanation:
                results[remark_col] = explanation
            else:
                results[remark_col] = "Based on available Oracle FLEXCUBE documentation and capabilities analysis, this specific requirement is not supported by the current platform architecture."
                
        elif response_value == "not found":
            # NOT FOUND: explanation only
            if explanation:
                results[remark_col] = explanation
            else:
                results[remark_col] = "Comprehensive analysis of available Oracle documentation and industry resources could not identify specific information regarding this requirement. Further clarification with Oracle technical teams may be required."
    else:
        # Fallback: use explanation only
        if remark_col and explanation:
            results[remark_col] = explanation

    return results

# ===============================
# Process Excel
# ===============================
@app.post("/process")
async def process_data(request: ProcessRequest):
    try:
        if not hasattr(app.state, 'temp_files') or request.filename not in app.state.temp_files:
            raise HTTPException(status_code=400, detail="File not found. Please upload again.")

        contents = app.state.temp_files[request.filename]
        df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        df.reset_index(drop=True, inplace=True)

        # Normalize column names
        df.columns = [col.strip().upper() for col in df.columns]
        input_cols = [col.strip().upper() for col in request.input_columns]
        output_cols = [col.strip().upper() for col in request.output_columns]

        # Add output columns if they don't exist
        for col in output_cols:
            if col not in df.columns:
                df[col] = ""

        # Enhanced System prompt with unbiased evaluation guidelines
        system_prompt = """
<assistantRole>
  <description>
    You are an expert AI assistant specializing in Oracle banking solutions and procurement analysis for the BFSI sector. You provide evidence-based, unbiased assessments of technical requirements against Oracle capabilities including FLEXCUBE, OFSAA, OBP, Digital Banking, and related Oracle banking technologies. Your responses are strictly based on available documentation and web search evidence.
  </description>

  <toolset>
    <tool>
      <name>Enhanced Web Search Analysis</name>
      <purpose>Comprehensive search across Oracle official documentation, support resources, industry publications, and technical communities to gather evidence about Oracle banking solution capabilities.</purpose>
    </tool>
    <usage>
      Utilize web search results to provide factual, evidence-based responses. Evaluate source quality and evidence strength to determine appropriate response confidence levels.
    </usage>
  </toolset>

  <guidelines strict="true">
    <decisive_evaluation>
      <instruction>Be decisive and evidence-based in your assessments. Avoid defaulting to "Partially" unless evidence specifically shows limited or conditional support.</instruction>
      
      <yes_criteria>
        Use "Yes" when:
        - Web search provides clear, specific evidence of full requirement support
        - Official Oracle documentation confirms the capability
        - Multiple reliable sources confirm complete functionality
        - Technical specifications clearly demonstrate requirement fulfillment
      </yes_criteria>
      
      <partially_criteria>
        Use "Partially" ONLY when:
        - Evidence shows the solution supports some but NOT ALL aspects of the requirement
        - Functionality exists but with documented limitations or conditions
        - Feature is available but requires additional components/licensing
        - Implementation is possible but with known constraints
      </partially_criteria>
      
      <no_criteria>
        Use "No" when:
        - Evidence clearly indicates the solution does NOT support the requirement
        - Official sources explicitly state feature unavailability
        - Multiple sources confirm lack of capability
        - Technical analysis shows incompatibility
      </no_criteria>
      
      <not_found_criteria>
        Use "Not found" when:
        - Comprehensive web search yields no relevant information
        - Available sources do not address the specific requirement
        - Evidence is insufficient to make any determination
      </not_found_criteria>
    </decisive_evaluation>

    <response_columns>
      <columns>TENDERER'S RESPONSE, VENDOR RESPONSE, ANSWER, COMPLIANCE</columns>
      <instruction>
        Provide decisive responses based on evidence strength:
        - Yes: Strong evidence of full support
        - Partially: Clear evidence of limited/conditional support  
        - No: Strong evidence of no support or incompatibility
        - Not found: Insufficient evidence for determination
        
        Evaluate each requirement independently. Do not assume partial support without specific evidence.
      </instruction>
    </response_columns>

    <remark_columns>
      <columns>TENDERER'S REMARK, VENDOR REMARKS, VENDOR COMMENTS, COMMENT, NOTES</columns>
      <instruction>
        Generate professional, descriptive explanations as if you are personally responding to the requirement. Write in a direct, confident, and professional manner explaining:
        - What the Oracle solution can/cannot do regarding the specific requirement
        - Technical capabilities, features, and implementation details
        - Version-specific information and architectural components when available
        - Integration approaches, APIs, and configuration options
        - Any limitations, prerequisites, or additional considerations
        - Professional assessment of compliance and implementation feasibility
        
        Write as a subject matter expert providing a comprehensive response. Use professional, confident language suitable for procurement documentation.
        
        Format: Write naturally and professionally, then add Reference Sources Consulted at the end with actual URLs.
        
        Example structure:
        "Oracle FLEXCUBE Universal Banking Solution provides comprehensive real-time transaction processing capabilities through its Real-Time Event Processing (RTEP) engine. The system supports instant balance updates across all banking channels including mobile banking, internet banking, ATM networks, and branch systems with sub-second response times. The architecture utilizes in-memory processing and event-driven design to ensure immediate transaction posting and balance visibility across all customer touchpoints.
        
        Reference Sources Consulted:
        Oracle Official Sources:
        • [Source Type] URL
        Industry & Technical Sources:  
        • [Source Type] URL"
        
        For insufficient information cases:
        "Based on comprehensive analysis of available Oracle documentation and industry resources, specific information regarding [requirement] could not be definitively established. While Oracle FLEXCUBE provides [general capabilities found], the particular aspects of [specific requirement] require further clarification with Oracle technical teams.
        
        Reference Sources Consulted:
        [List of URLs searched]"
        
        CRITICAL: Be professional, descriptive, and confident. Avoid technical jargon explanations, verbose analysis sections, or meta-commentary about the search process. Write as an expert providing a direct assessment.
      </instruction>
    </remark_columns>

    <explanation_requirement>
      <instruction>
        Generate professional, descriptive explanations written as expert responses justifying the assessment. Write naturally and confidently as a subject matter expert providing direct answers. Include:
        - Clear reasoning for the Yes/No/Partially/Not found determination
        - Technical context and implementation details discovered
        - Professional assessment of the Oracle solution's capabilities
        - Any relevant architectural, functional, or integration considerations
        
        Format: Write as a professional consultant providing a direct, confident response (3-4 sentences).
        
        Example: "Oracle FLEXCUBE Universal Banking Solution fully supports real-time transaction processing through its Real-Time Event Processing (RTEP) engine, delivering sub-second response times across all banking channels. The system's in-memory processing architecture ensures immediate balance updates across mobile banking, internet banking, ATM networks, and branch systems. This capability is integral to the core banking platform and requires no additional licensing or third-party components for standard banking operations. Implementation typically involves API configuration and channel integration setup through Oracle's standard deployment methodology."
        
        Avoid: Meta-commentary about evidence strength, search processes, or confidence levels. Write as an expert providing direct assessment.
      </instruction>
    </explanation_requirement>

    <evidence_evaluation>
      <high_confidence>Multiple Oracle official sources, clear technical documentation, confirmed capabilities</high_confidence>
      <moderate_confidence>Some official sources, industry confirmation, reasonable technical evidence</moderate_confidence>
      <limited_confidence>Few sources, indirect evidence, technical discussion without official confirmation</limited_confidence>
      <no_confidence>No relevant sources, no technical evidence, comprehensive search yielded no results</no_confidence>
    </evidence_evaluation>

    <strict_rules>
      <rule>Base all responses strictly on web search evidence - never assume or extrapolate beyond available data</rule>
      <rule>Evaluate each requirement independently without bias toward any particular response</rule>
      <rule>Provide specific technical details and version information when available</rule>
      <rule>Reference source quality and evidence strength in assessments</rule>
      <rule>Be decisive when evidence clearly supports Yes or No determinations</rule>
      <rule>Use professional, procurement-appropriate language throughout</rule>
      <rule>Include comprehensive source analysis in remarks</rule>
    </strict_rules>

    <output_format>
      <instruction>
        Structure responses exactly as follows:
        [COLUMN_NAME]: [Response Value]
        
        EXPLANATION: [Detailed professional explanation with evidence justification]
      </instruction>
    </output_format>
  </guidelines>
</assistantRole>
"""

        async def process_row_async(index: int, row: pd.Series) -> tuple[int, Dict[str, str]]:
            """Enhanced row processing with comprehensive evidence evaluation"""
            start_time = time.time()
            
            try:
                # Extract input data efficiently
                input_data = {}
                for col in input_cols:
                    if col in row and pd.notna(row[col]):
                        val = str(row[col]).strip()
                        if val:
                            input_data[col] = val
                
                input_text = " ".join(input_data.values())
                word_count = len(input_text.split())

                if word_count < 5:
                    logger.warning(f"Row {index + 1}: Skipped - insufficient content for analysis")
                    return index, {col: "Insufficient content for analysis" for col in output_cols}

                # Create focused, comprehensive search query
                key_terms = []
                for val in input_data.values():
                    # Extract key technical terms and concepts
                    words = [w for w in val.split() if len(w) > 3][:150]
                    key_terms.extend(words)
                
                # Enhanced search query construction
                search_query = f"oracle flexcube {' '.join(key_terms[:100])}"
                
                # Get comprehensive search results with source tracking
                search_results = await exa_search(search_query, row_index=index)
                
                # Process search content and metadata
                search_content = search_results.get("content", "") if isinstance(search_results, dict) else str(search_results)
                sources = search_results.get("sources", []) if isinstance(search_results, dict) else []
                source_types = search_results.get("source_types", []) if isinstance(search_results, dict) else []
                evidence_strength = search_results.get("evidence_strength", "None") if isinstance(search_results, dict) else "None"
                oracle_sources = search_results.get("oracle_sources", 0) if isinstance(search_results, dict) else 0
                community_sources = search_results.get("community_sources", 0) if isinstance(search_results, dict) else 0
                
                # Optimize content length while preserving quality
                if len(search_content) > 4000:
                    search_content = search_content[:4000] + "... (content truncated for processing efficiency while maintaining key technical details)"

                # Create comprehensive input prompt
                input_text_prompt = "\n".join([f"{k}: {v[:300]}" for k, v in input_data.items()])
                
                # Enhanced source context for AI analysis
                source_context = ""
                if sources:
                    unique_source_types = list(set(source_types))
                    source_context = f"""
Source Analysis Context:
- Total sources found: {len(sources)}
- Oracle official sources: {oracle_sources}
- Community/Industry sources: {community_sources}
- Evidence strength: {evidence_strength}
- Source categories: {', '.join(unique_source_types)}
- Assessment guidance: {'Use decisive Yes/No responses with high confidence' if evidence_strength == 'High' else 'Evaluate carefully based on available evidence' if evidence_strength == 'Moderate' else 'Consider Partially/Not found responses due to limited evidence' if evidence_strength == 'Limited' else 'Likely Not found response due to no relevant evidence'}
"""
                
                # Enhanced prompt with response-specific formatting requirements
                full_prompt = f"""{system_prompt}

User Specific Instructions: {request.user_prompt}

Excel Input Requirement:
{input_text_prompt}

Web Search Results Analysis:
{search_content}
{source_context}

Required Output Columns: {', '.join(output_cols)}

CRITICAL FORMATTING REQUIREMENTS BASED ON RESPONSE:

IF RESPONSE IS "YES":
- Remark column: Detailed professional explanation of how Oracle FLEXCUBE supports the requirement + Reference Sources Consulted section with URLs

IF RESPONSE IS "PARTIALLY": 
- Remark column: Detailed professional explanation of what Oracle FLEXCUBE supports and what limitations exist + Reference Sources Consulted section with URLs

IF RESPONSE IS "NO":
- Remark column: Professional explanation of why the requirement is not supported, mention that specific capability could not be found in Oracle FLEXCUBE documentation (NO Reference Sources Consulted section)

IF RESPONSE IS "NOT FOUND":
- Remark column: Professional explanation that comprehensive analysis could not identify information about this requirement (NO Reference Sources Consulted section)

RESPONSE FORMAT EXAMPLE FOR YES/PARTIALLY:
[REMARK COLUMN]: Oracle FLEXCUBE Universal Banking Solution provides comprehensive real-time transaction processing through its RTEP engine with sub-second response times across all banking channels...

Reference Sources Consulted:
Oracle Official Sources:
• [Source Type] URL
Industry & Technical Sources:  
• [Source Type] URL

RESPONSE FORMAT EXAMPLE FOR NO/NOT FOUND:
[REMARK COLUMN]: Based on available Oracle FLEXCUBE documentation and capabilities analysis, this specific requirement is not supported by the current platform architecture...
(NO Reference Sources Consulted section)

EXPLANATION: [3-4 sentence professional justification]

CRITICAL: Only include "Reference Sources Consulted" for YES and PARTIALLY responses. For NO and NOT FOUND responses, provide explanation only.
"""

                messages = [
                    {"role": "system", "content": "You are an expert Oracle banking solutions analyst. Provide evidence-based, unbiased assessments. Be decisive when evidence clearly supports Yes or No responses. Avoid defaulting to 'Partially' unless evidence specifically indicates limited functionality."},
                    {"role": "user", "content": full_prompt}
                ]
                
                logger.info("\n🔄 SENDING TO AI ANALYSIS")
                logger.info(f"{'='*100}")
                logger.info(f"🎯 Row: {index + 1}")
                logger.info(f"📊 Evidence Strength: {evidence_strength}")
                logger.info(f"🏢 Oracle Sources: {oracle_sources}")
                logger.info(f"🌐 Community Sources: {community_sources}")
                logger.info(f"📚 Source Categories: {len(set(source_types))}")
                logger.info(f"🤖 Model: {MODEL_NAME} | Max Tokens: 1200")
                logger.info(f"{'='*100}")
                
                ai_response = await call_openai(messages, max_tokens=1200)
                
                logger.info("\n🎉 AI RESPONSE RECEIVED:")
                logger.info(f"{'='*80}")
                logger.info(f"{ai_response}")
                logger.info(f"{'='*80}")
                
                # Enhanced extraction with comprehensive source information
                search_info = {
                    "sources": sources,
                    "source_types": source_types,
                    "evidence_strength": evidence_strength,
                    "oracle_sources": oracle_sources,
                    "community_sources": community_sources
                }
                result = extract_column_values(ai_response, output_cols, search_info)
                
                elapsed = time.time() - start_time
                
                logger.info(f"\n✅ ROW {index + 1} - PROCESSING COMPLETED")
                logger.info(f"{'='*120}")
                logger.info(f"⏱️  Processing Time: {elapsed:.2f} seconds")
                logger.info(f"🔍 Total Sources Analyzed: {len(sources)}")
                logger.info(f"🏆 Evidence Quality: {evidence_strength}")
                logger.info("📊 FINAL EXTRACTED VALUES:")
                for col, val in result.items():
                    logger.info(f"\n   📋 {col}:")
                    # Improved display formatting
                    val_preview = val[:300] + "..." if len(val) > 300 else val
                    logger.info(f"      {textwrap.fill(val_preview, width=95, subsequent_indent='      ')}")
                logger.info(f"\n🎯 Assessment Quality: {'High confidence' if evidence_strength == 'High' else 'Moderate confidence' if evidence_strength == 'Moderate' else 'Limited confidence' if evidence_strength == 'Limited' else 'Insufficient evidence'}")
                logger.info(f"{'='*120}")
                
                return index, result
                
            except Exception as e:
                elapsed = time.time() - start_time
                logger.error(f"\n❌ ROW {index + 1} - PROCESSING ERROR")
                logger.error(f"{'='*100}")
                logger.error(f"⏱️  Time Elapsed: {elapsed:.2f} seconds")
                logger.error(f"🚨 Error Details: {str(e)}")
                logger.error(f"📝 Input Preview: {input_text[:200]}...")
                logger.error(f"{'='*100}")
                logger.exception(f"Processing error for row {index + 1}")
                
                error_result = {col: f"Processing error: {str(e)[:150]}..." for col in output_cols}
                return index, error_result

        async def process_all_rows():
            """Enhanced batch processing with comprehensive progress tracking"""
            overall_start = time.time()
            total_rows = len(df)
            
            logger.info(f"Starting enhanced processing of {total_rows} rows")
            
            # Processing statistics
            stats = {
                "high_evidence": 0,
                "moderate_evidence": 0, 
                "limited_evidence": 0,
                "no_evidence": 0,
                "total_oracle_sources": 0,
                "total_community_sources": 0,
                "cache_hits": 0
            }
            
            for batch_start in range(0, total_rows, BATCH_SIZE):
                batch_end = min(batch_start + BATCH_SIZE, total_rows)
                batch_rows = list(range(batch_start, batch_end))
                
                batch_tasks = []
                for idx in batch_rows:
                    task = process_row_async(idx, df.iloc[idx])
                    batch_tasks.append(task)
                
                batch_start_time = time.time()
                batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)
                batch_elapsed = time.time() - batch_start_time
                
                # Process batch results and update statistics
                for result in batch_results:
                    if isinstance(result, Exception):
                        logger.error(f"\n❌ BATCH ERROR: {result}")
                        logger.exception(f"Batch processing error")
                        continue
                    
                    idx, row_results = result
                    for col, val in row_results.items():
                        df.at[idx, col] = val
                    
                    print(f"\n💾 ROW {idx + 1} - EXCEL DATA UPDATED")
                    print(f"{'='*90}")
                    print("📊 FINAL COLUMN VALUES IN EXCEL:")
                    for col in output_cols:
                        excel_value = df.at[idx, col]
                        val_display = str(excel_value)[:200] + "..." if len(str(excel_value)) > 200 else str(excel_value)
                        print(f"\n   📋 {col}:")
                        print(f"      {textwrap.fill(val_display, width=80, subsequent_indent='      ')}")
                    print(f"{'='*90}")
                
                completed = batch_end
                progress = (completed / total_rows) * 100
                
                print("\n🚀 BATCH PROGRESS UPDATE")
                print(f"{'='*80}")
                print(f"✅ Completed: {completed}/{total_rows} rows ({progress:.1f}%)")
                print(f"⏱️  Batch Time: {batch_elapsed:.2f}s | Avg per Row: {batch_elapsed/len(batch_rows):.2f}s")
                print(f"💾 Cache Size: {len(search_cache)} entries")
                print("🎯 Processing Quality: Enhanced evidence-based analysis active")
                print(f"{'='*80}")
                
                if batch_end < total_rows:
                    print(f"\n⏸️  Brief pause before next batch...")
                    await asyncio.sleep(0.5)

            total_elapsed = time.time() - overall_start
            
            logger.info("\n🎊 ENHANCED PROCESSING COMPLETED!")
            logger.info(f"{'='*120}")
            logger.info("📊 COMPREHENSIVE FINAL STATISTICS:")
            logger.info(f"   ✅ Total Rows Processed: {total_rows}")
            logger.info(f"   ⏱️  Total Processing Time: {total_elapsed:.2f} seconds")
            logger.info(f"   📈 Average Time per Row: {total_elapsed/total_rows:.2f} seconds")
            logger.info(f"   💾 Final Cache Size: {len(search_cache)} entries")
            
            # Analyze cache for quality metrics
            cache_with_sources = sum(1 for v in search_cache.values() if isinstance(v, dict) and len(v.get('sources', [])) > 0)
            cache_hit_rate = (cache_with_sources / max(1, len(search_cache))) * 100
            
            logger.info(f"   🎯 Cache Quality: {cache_hit_rate:.1f}% entries with sources")
            logger.info("   🔍 Search Effectiveness: Enhanced source tracking and evidence analysis")
            logger.info("   🏆 Assessment Quality: Evidence-based evaluation with bias reduction")
            logger.info("   📚 Source Integration: Comprehensive source type analysis and referencing")
            logger.info(f"{'='*120}")

        # Start enhanced processing
        logger.info("\n🚀 STARTING ENHANCED EXCEL PROCESSING")
        logger.info(f"{'='*120}")
        logger.info("📊 Enhanced Processing Parameters:")
        logger.info(f"   📝 Total Rows: {len(df)}")
        logger.info(f"   📥 Input Columns: {', '.join(input_cols)}")
        logger.info(f"   📤 Output Columns: {', '.join(output_cols)}")
        logger.info(f"   🔄 Batch Size: {BATCH_SIZE}")
        logger.info(f"   🤖 AI Model: {MODEL_NAME} (Enhanced prompting)")
        logger.info(f"   💾 Max Cache Size: {CACHE_SIZE}")
        logger.info("   🎯 Key Enhancements:")
        logger.info("      • Unbiased response evaluation")
        logger.info("      • Comprehensive source tracking")
        logger.info("      • Evidence strength assessment")
        logger.info("      • Professional remark generation")
        logger.info("      • Enhanced explanation quality")
        logger.info(f"{'='*120}")
        
        await process_all_rows()

        # Generate Excel file with enhanced formatting
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Enhanced Processed Data')
            worksheet = writer.sheets['Enhanced Processed Data']

            # Enhanced styling
            header_font = Font(bold=True, color='FFFFFF', size=12)
            header_fill = PatternFill("solid", fgColor="2F5597")

            for col_num, column_title in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                header_cell = worksheet[f"{col_letter}1"]
                header_cell.font = header_font
                header_cell.fill = header_fill
                
                # Enhanced column width based on content type
                if any(keyword in column_title.upper() for keyword in ["REMARK", "COMMENT", "NOTES"]):
                    worksheet.column_dimensions[col_letter].width = 80  # Wider for detailed remarks
                elif any(keyword in column_title.upper() for keyword in ["RESPONSE", "ANSWER", "COMPLIANCE"]):
                    worksheet.column_dimensions[col_letter].width = 25
                else:
                    worksheet.column_dimensions[col_letter].width = 30

            # Enhanced row formatting
            max_display_rows = min(1000, worksheet.max_row)
            for row_idx in range(2, max_display_rows + 1):
                worksheet.row_dimensions[row_idx].height = 80  # Increased for detailed content
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

        output.seek(0)
        file_id = str(uuid.uuid4())
        if not hasattr(app.state, 'processed_files'):
            app.state.processed_files = {}
        app.state.processed_files[file_id] = output.getvalue()
        
        # Cleanup temporary files
        if request.filename in app.state.temp_files:
            del app.state.temp_files[request.filename]

        return {
            "file_id": file_id, 
            "message": f"✅ Enhanced processing completed successfully - {len(df)} rows analyzed with comprehensive evidence-based evaluation",
            "processing_stats": {
                "total_rows": len(df),
                "cache_entries": len(search_cache),
                "enhancement_features": [
                    "Unbiased response evaluation",
                    "Comprehensive source tracking", 
                    "Evidence strength assessment",
                    "Professional remark generation",
                    "Enhanced explanation quality"
                ]
            },
            "processing_complete": True
        }

    except Exception as e:
        logger.exception(f"Enhanced processing error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Enhanced processing error: {str(e)}")

# ===============================
# Download Processed Excel
# ===============================
@app.get("/download/{file_id}")
async def download_file(file_id: str):
    try:
        if not hasattr(app.state, 'processed_files') or file_id not in app.state.processed_files:
            raise HTTPException(status_code=404, detail="File not found or expired")

        return StreamingResponse(
            io.BytesIO(app.state.processed_files[file_id]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=enhanced_processed_requirements.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Download error: {str(e)}")

# ===============================
# Enhanced Utility Endpoints
# ===============================
@app.get("/health")
async def health_check():
    return {
        "status": "healthy", 
        "timestamp": datetime.now().isoformat(),
        "cache_size": len(search_cache),
        "model": MODEL_NAME,
        "enhancements": "Evidence-based analysis with comprehensive source tracking"
    }

@app.get("/clear-cache")
async def clear_cache():
    """Clear the enhanced search cache"""
    global search_cache
    cache_size = len(search_cache)
    
    # Analyze cache before clearing
    sources_count = sum(len(v.get('sources', [])) for v in search_cache.values() if isinstance(v, dict))
    oracle_sources = sum(v.get('oracle_sources', 0) for v in search_cache.values() if isinstance(v, dict))
    
    search_cache.clear()
    return {
        "message": "Enhanced cache cleared successfully", 
        "details": {
            "entries_removed": cache_size,
            "total_sources_cleared": sources_count,
            "oracle_sources_cleared": oracle_sources
        }
    }

@app.get("/cache-stats")
async def get_cache_stats():
    """Get detailed cache statistics"""
    if not search_cache:
        return {"message": "Cache is empty"}
    
    total_entries = len(search_cache)
    entries_with_sources = sum(1 for v in search_cache.values() if isinstance(v, dict) and len(v.get('sources', [])) > 0)
    total_sources = sum(len(v.get('sources', [])) for v in search_cache.values() if isinstance(v, dict))
    oracle_sources = sum(v.get('oracle_sources', 0) for v in search_cache.values() if isinstance(v, dict))
    community_sources = sum(v.get('community_sources', 0) for v in search_cache.values() if isinstance(v, dict))
    
    evidence_strength_counts = {}
    for v in search_cache.values():
        if isinstance(v, dict):
            strength = v.get('evidence_strength', 'Unknown')
            evidence_strength_counts[strength] = evidence_strength_counts.get(strength, 0) + 1
    
    return {
        "cache_overview": {
            "total_entries": total_entries,
            "entries_with_sources": entries_with_sources,
            "success_rate": f"{(entries_with_sources/total_entries)*100:.1f}%" if total_entries > 0 else "0%"
        },
        "source_analysis": {
            "total_sources": total_sources,
            "oracle_sources": oracle_sources,
            "community_sources": community_sources,
            "average_sources_per_entry": f"{total_sources/max(1, entries_with_sources):.1f}"
        },
        "evidence_strength_distribution": evidence_strength_counts
    }

@app.get("/stats")
async def get_stats():
    """Get comprehensive processing statistics"""
    return {
        "system_config": {
            "cache_entries": len(search_cache),
            "max_concurrent_requests": MAX_CONCURRENT_REQUESTS,
            "batch_size": BATCH_SIZE,
            "model": MODEL_NAME,
            "rate_limit_delay": RATE_LIMIT_DELAY
        },
        "enhancements": {
            "unbiased_evaluation": "Active",
            "source_tracking": "Comprehensive",
            "evidence_assessment": "Multi-level",
            "professional_remarks": "Enhanced",
            "explanation_quality": "Detailed"
        },
        "performance": {
            "cache_size_limit": CACHE_SIZE,
            "concurrent_processing": True,
            "batch_optimization": True
        }
    }

@app.get("/")
async def root():
    return {
        "message": "Enhanced AI Excel Processor API - Evidence-Based Analysis", 
        "version": "4.0 Enhanced", 
        "key_improvements": [
            "🎯 Unbiased response evaluation (eliminates 'Partially' bias)",
            "🔍 Comprehensive source tracking and analysis",
            "📊 Evidence strength assessment (High/Moderate/Limited)",
            "📚 Professional remark generation with source context",
            "📝 Enhanced explanation quality with technical depth",
            "🏢 Oracle vs community source differentiation",
            "⚡ Improved processing efficiency"
        ],
        "api_features": {
            "max_concurrent": MAX_CONCURRENT_REQUESTS,
            "batch_size": BATCH_SIZE,
            "evidence_based_analysis": True,
            "source_categorization": True,
            "professional_output": True
        }
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        app, 
        host="0.0.0.0", 
        port=8000,
        log_level="info",
        access_log=False
    )